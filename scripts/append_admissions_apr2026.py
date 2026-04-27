"""2026-04 入院データ (1〜26 日) を既存 past_admissions_2025fy.csv に統合.

副院長要望 (2026-04-27):
    事務から `20260427(0401-0426).xlsx` (137 件) 受領。これを既存 1 年分
    (2025-04-01〜2026-03-31、1,823 件) に追記し、2025FY+2026Apr 統合
    データとして看護必要度・救急15%・LOS 等の rolling 計算に反映する。

特に注目: 4 月から強化したペイン科キシロカイン点滴 (A6⑧ 抗不整脈) が
看護必要度 A 項目に効果を発揮したかの追跡。

入力:
    /Users/torukubota/Downloads/20260427(0401-0426).xlsx (137 件)

出力:
    data/past_admissions_2025fy.csv (上書き、合計 1,960 件想定)

CLI:
    .venv/bin/python scripts/append_admissions_apr2026.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd


XLSX_PATH = Path("/Users/torukubota/Downloads/20260427(0401-0426).xlsx")
CSV_PATH = Path(__file__).resolve().parent.parent / "data" / "past_admissions_2025fy.csv"

# 病棟全角→半角の変換
WARD_MAP = {"６階": "6F", "５階": "5F"}

# 医師実名→コード辞書 (既存 CSV のコード命名に準拠、新規 2 名追加)
DOCTOR_NAME_TO_CODE = {
    "照屋　宏充": "TERUH",
    "近藤　章之": "KONA",
    "井上　卓": "INOT",
    "久保田　徹": "KUBT",
    "加治佐　淳一": "KJJ",
    "林　貴徳": "HAYT",
    "奥田　和弘": "OKUK",
    "城間　健治": "SIROK",
    "大城　譲": "OHSY",
    "外間　政朗": "HOKM",
    "平良　勝己": "TAIRK",
    "比嘉　昇": "HIGN",
    "比嘉　達也": "HIGT",
    "福田　昌輝": "FKDM",
    # 新規 (2026-04 で初登場)
    "森　直樹": "MORN",
    "健山　正男": "TAKY",
}


def _normalize_date(v) -> str:
    """様々な形式の日付を YYYY-MM-DD に統一."""
    if v is None or v == "":
        return ""
    if isinstance(v, str):
        return v.replace("/", "-")
    # datetime
    return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"❌ XLSX が見つかりません: {XLSX_PATH}")
        return 1
    if not CSV_PATH.exists():
        print(f"❌ 既存 CSV が見つかりません: {CSV_PATH}")
        return 1

    # 既存 CSV 読み込み
    existing = pd.read_csv(CSV_PATH)
    print(f"📊 既存 CSV: {len(existing)} 件 ({existing['入院日'].min()} 〜 {existing['入院日'].max()})")

    # 新 xlsx 読み込み
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb["Sheet1"]
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    new_records: list[dict] = []
    next_id = int(existing["患者番号"].max()) + 1
    unknown_doctors: set[str] = set()

    for r in rows:
        d = dict(zip(headers, r))
        # 列マッピング
        ward = WARD_MAP.get(d.get("病棟", ""), d.get("病棟", ""))
        doctor_name = d.get("医師", "")
        doctor_code = DOCTOR_NAME_TO_CODE.get(doctor_name, "")
        if not doctor_code and doctor_name:
            unknown_doctors.add(doctor_name)
            doctor_code = "UNKNOWN"

        new_records.append({
            "患者番号": next_id,
            "病棟": ward,
            "救急車": d.get("救急車", ""),
            "緊急": d.get("緊急", ""),
            "入経路": d.get("入経路", ""),
            "入院日": _normalize_date(d.get("入院日")),
            "退院日": _normalize_date(d.get("退院日")),
            "退経路": d.get("退経路", ""),
            "日数": d.get("日数", 0),
            "診療科": d.get("診療科", ""),
            "医師": doctor_code,
            "手術": d.get("手術", ""),
        })
        next_id += 1

    new_df = pd.DataFrame(new_records, columns=existing.columns.tolist())

    if unknown_doctors:
        print(f"⚠️ 未知の医師名 (UNKNOWN として保存): {unknown_doctors}")

    print(f"📥 新規 (xlsx): {len(new_df)} 件 ({new_df['入院日'].min()} 〜 {new_df['入院日'].max()})")

    # 結合
    combined = pd.concat([existing, new_df], ignore_index=True)
    print(f"📦 結合後: {len(combined)} 件")

    # 病棟分布
    print(f"   病棟分布: {combined['病棟'].value_counts().to_dict()}")
    print(f"   2026-04 ペイン科 (麻酔科) 件数: {(new_df['診療科'] == '麻酔科').sum()}")
    print(f"   2026-04 救急車入院: {(new_df['救急車'] == '有り').sum()}")

    # 上書き保存
    combined.to_csv(CSV_PATH, index=False)
    print(f"✅ 保存: {CSV_PATH}")
    print(f"   期間: {combined['入院日'].min()} 〜 {combined['入院日'].max()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
