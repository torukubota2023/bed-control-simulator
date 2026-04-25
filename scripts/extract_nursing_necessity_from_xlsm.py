"""看護必要度 XLSM → CSV 抽出スクリプト.

事務提供の看護必要度データシート（月次 XLSM、12 ヶ月分）から「データ（全体）」
シートのみを抽出し、`data/nursing_necessity_2025fy.csv` に集計データを書き出す。

**個人情報保護**:
- 「データ（全体）」シート（日次×病棟集計、個人情報なし）のみ取り込む
- 「患者別」「MedFile」シート（個人情報あり）は取り込まない
- 元 XLSM はリポジトリ外（`/tmp` 等）に保管、本スクリプトでは CSV のみ生成

使用方法:
    # 副院長月次運用例
    .venv/bin/python scripts/extract_nursing_necessity_from_xlsm.py \\
        --src /tmp/nursing_2025/看護必要度2025/ \\
        --dest data/nursing_necessity_2025fy.csv

入力ファイル名パターン:
    看護必要度データシート_<YEAR>年<MONTH>月_EF統合ファイル反映.xlsm
"""

from __future__ import annotations

import argparse
import glob
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd


# 「データ（全体）」シートの列インデックス（1-based を 0-based に）
COL_DATE = 1            # C2: 日付
COL_WARD = 2            # C3: 病棟
COL_TEISHO = 4          # C5: 定床数
COL_I_TOTAL = 28        # C29: 必要度Ⅰ 地域包括 患者延べ数
COL_I_PASS1 = 29        # C30: 必要度Ⅰ 地域包括 基準①超対象患者数
COL_I_RATE1 = 30        # C31: 必要度Ⅰ 地域包括 基準①「満たす割合」
COL_I_ADMIT = 31        # C32: 必要度Ⅰ 地域包括 入棟患者数
COL_I_PASS2 = 32        # C33: 必要度Ⅰ 地域包括 基準②超対象患者数
COL_I_RATE2 = 33        # C34: 必要度Ⅰ 地域包括 基準②「満たす割合」
COL_II_TOTAL = 57       # C58: 必要度Ⅱ 地域包括 患者延べ数
COL_II_PASS1 = 58       # C59: 必要度Ⅱ 地域包括 基準①超対象患者数
COL_II_RATE1 = 59       # C60: 必要度Ⅱ 地域包括 基準①「満たす割合」
COL_II_ADMIT = 60       # C61: 必要度Ⅱ 地域包括 入棟患者数
COL_II_PASS2 = 61       # C62: 必要度Ⅱ 地域包括 基準②超対象患者数
COL_II_RATE2 = 62       # C63: 必要度Ⅱ 地域包括 基準②「満たす割合」


def extract_one_xlsm(filepath: str) -> list[dict]:
    """1 つの XLSM から「データ（全体）」シートを抽出.

    Returns:
        list of dict (rows)。日付パース失敗・病棟識別不可の行はスキップ。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if "データ（全体）" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{filepath} に「データ（全体）」シートが見つかりません")

    ws = wb["データ（全体）"]
    rows: list[dict] = []

    # データ行は R8 以降（R6-R7 はヘッダー）
    for row in ws.iter_rows(min_row=8, values_only=True):
        date_val = row[COL_DATE]
        ward = row[COL_WARD]
        if date_val is None:
            continue

        date_str = str(date_val)
        # 「【合計】2025/04/01」形式と「2025/04/01」形式の両方に対応
        if date_str.startswith("【合計】"):
            ward_label = "合計"
            date_str = date_str.replace("【合計】", "")
        elif ward and "５Ｆ" in str(ward):
            ward_label = "5F"
        elif ward and "６Ｆ" in str(ward):
            ward_label = "6F"
        else:
            continue

        try:
            d = datetime.strptime(date_str.strip(), "%Y/%m/%d").date()
        except ValueError:
            continue

        rows.append({
            "date": d.isoformat(),
            "ward": ward_label,
            "teisho": row[COL_TEISHO],
            "I_total": row[COL_I_TOTAL],
            "I_pass1": row[COL_I_PASS1],
            "I_rate1": row[COL_I_RATE1],
            "I_admit": row[COL_I_ADMIT],
            "I_pass2": row[COL_I_PASS2],
            "I_rate2": row[COL_I_RATE2],
            "II_total": row[COL_II_TOTAL],
            "II_pass1": row[COL_II_PASS1],
            "II_rate1": row[COL_II_RATE1],
            "II_admit": row[COL_II_ADMIT],
            "II_pass2": row[COL_II_PASS2],
            "II_rate2": row[COL_II_RATE2],
        })
    wb.close()
    return rows


def extract_directory(src_dir: str, dest_csv: str) -> int:
    """ディレクトリ内の全 XLSM を統合し CSV に書き出す.

    Returns:
        書き出した行数
    """
    files = sorted(glob.glob(f"{src_dir}/*.xlsm"))
    if not files:
        print(f"⚠️ {src_dir} に .xlsm ファイルがありません", file=sys.stderr)
        return 0

    print(f"📁 検出ファイル: {len(files)} 件")
    all_rows: list[dict] = []
    for fp in files:
        try:
            rows = extract_one_xlsm(fp)
            print(f"  ✅ {Path(fp).name}: {len(rows)} 行")
            all_rows.extend(rows)
        except Exception as e:
            print(f"  ❌ {Path(fp).name}: {e}", file=sys.stderr)

    if not all_rows:
        print("⚠️ 抽出できたデータがありません", file=sys.stderr)
        return 0

    df = pd.DataFrame(all_rows)
    # date でソートして再現性を確保
    df = df.sort_values(["date", "ward"]).reset_index(drop=True)

    Path(dest_csv).parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(dest_csv, index=False)
    print(f"\n💾 書き出し完了: {dest_csv}  ({len(df)} 行)")
    print(f"期間: {df['date'].min()} ～ {df['date'].max()}")
    print(f"病棟: {df['ward'].value_counts().to_dict()}")
    return len(df)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="看護必要度 XLSM → CSV 抽出"
    )
    parser.add_argument(
        "--src",
        required=True,
        help="XLSM が格納されているディレクトリ（例: /tmp/nursing_2025/看護必要度2025/）",
    )
    parser.add_argument(
        "--dest",
        default="data/nursing_necessity_2025fy.csv",
        help="出力 CSV パス（デフォルト: data/nursing_necessity_2025fy.csv）",
    )
    args = parser.parse_args()

    n = extract_directory(args.src, args.dest)
    return 0 if n > 0 else 1


if __name__ == "__main__":
    sys.exit(main())
